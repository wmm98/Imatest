from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from Conf.config import Config
from Common.interface import Interface
from Common.color_data_filter import CSVTestData
from Common.get_report_position import GetReportPosition
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side

conf = Config()


class WriteReport(Interface):
    def __init__(self, file_path, template_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.template_path = template_path
        self.red_font = Font(color="FF0000")
        self.is_quality = False
        self.scenario_light = ""
        self.yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        # 问题总结数据准备
        self.questions_summary = []
        self.F_light_description = []
        self.D65_light_description = []
        self.CWF_light_description = []
        self.TL84_light_description = []
        self.HJ_light_description = []
        self.report_position = GetReportPosition(self.template_path, conf.sheet_name)

    def get_border(self):
        # 创建一个细边框样式
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        return thin_border

    def write_questions_summary_data(self):
        wb = load_workbook(self.file_path)
        sheet = wb[conf.summary_sheet_name]
        # 写入总结
        # 找到位置
        if len(self.questions_summary) != 0:
            key_position = GetReportPosition(self.template_path, conf.summary_sheet_name).find_scenario_position_by_keyword(conf.r_summary_description)
            for i in range(1, len(self.questions_summary) + 1):
                if len(self.questions_summary[i - 1]) != 0:
                    queue_pos_row = key_position[0] + i
                    queue_pos_clo = key_position[0] - 1
                    summary_pos_row = key_position[0] + i
                    summary_pos_clo = key_position[1]
                    detail_pos_row = key_position[0] + i
                    detail_pos_clo = key_position[0] + 1

                    queue_cell = sheet.cell(row=queue_pos_row, column=queue_pos_clo)
                    summary_cell = sheet.cell(row=summary_pos_row, column=summary_pos_clo)
                    detail_cell = sheet.cell(row=detail_pos_row, column=detail_pos_clo)

                    # 设置行高
                    sheet.row_dimensions[queue_pos_row].height = 30  # 设置行高为30个单位
                    queue_cell.value = i
                    queue_cell.alignment = Alignment(horizontal='center', vertical='center')
                    summary_cell.value = self.questions_summary[i - 1][0] + "，".join(self.questions_summary[i - 1][1:])
                    detail_cell.value = self.questions_summary[i - 1][0] + "，".join(self.questions_summary[i - 1][1:])

                    # 给cell 加上边框
                    # 最后一列, 给表格加上边框
                    last_clo = key_position[1] + 6
                    for c in range(key_position[1] - 1, last_clo + 1):
                        e_cell = sheet.cell(row=queue_pos_row, column=c)
                        e_cell.border = self.get_border()
                        # 第一列不用设置，上面已经设置好
                        if c != key_position[1] - 1:
                            e_cell.alignment = Alignment(vertical='center')

        wb.save(self.file_path)
        wb.close()

    def write_contrast_data(self, position, value):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        pos = position[conf.r_hj_relate_pos]
        contrast_cell = sheet.cell(row=pos[0], column=pos[1])
        contrast_cell.value = "%s%%" % ("{:.2f}".format(value[conf.hj_contrast] * 100))

        self.comparative_contrast_indicator(contrast_cell, value[conf.hj_contrast])
        wb.save(self.file_path)
        wb.close()

    def write_readable_hj_data(self, position, value):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        pos = position[conf.r_hj_relate_pos]
        readable_cell = sheet.cell(row=pos[0], column=pos[1])
        readable_cell.value = value[conf.hj_readable]

        self.comparative_readable_hj_indicator(readable_cell, value[conf.hj_readable])
        wb.save(self.file_path)
        wb.close()

    def write_dynamic_range_data(self, position, value):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        # 写入灰阶动态范围
        pos = position[conf.r_hj_relate_pos]
        dynamic_range_cell = sheet.cell(row=pos[0], column=pos[1])
        dynamic_range_cell.value = value[conf.hj_dynamic_range]

        self.comparative_dynamic_range_indicator(dynamic_range_cell, value[conf.hj_dynamic_range])
        wb.save(self.file_path)
        wb.close()

    def write_snr_data(self, position, value):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]

        r_cell = sheet.cell(row=position[conf.R][0], column=position[conf.R][1])
        r_cell.value = "R：%s" % str(value[conf.R])
        g_cell = sheet.cell(row=position[conf.G][0], column=position[conf.G][1])
        g_cell.value = "G：%s" % str(value[conf.G])
        b_cell = sheet.cell(row=position[conf.B][0], column=position[conf.B][1])
        b_cell.value = "B：%s" % str(value[conf.B])
        y_cell = sheet.cell(row=position[conf.Y][0], column=position[conf.Y][1])
        y_cell.value = "Y：%s" % str(value[conf.Y])

        self.comparative_snr_indicator(
            [[r_cell, value[conf.R]], [g_cell, value[conf.G]], [b_cell, value[conf.B]], [y_cell, value[conf.Y]]])

        wb.save(self.file_path)
        wb.close()

    def write_white_balance_data(self, position, value):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]

        nineteen_cell = sheet.cell(row=position[conf.nineteen][0], column=position[conf.nineteen][1])
        nineteen_cell.value = "%s block： %s" % (conf.nineteen, str(value[conf.nineteen]))
        twenty_cell = sheet.cell(row=position[conf.twenty][0], column=position[conf.twenty][1])
        twenty_cell.value = "%s block： %s" % (conf.twenty, str(value[conf.twenty]))
        twenty_one_cell = sheet.cell(row=position[conf.twenty_one][0], column=position[conf.twenty_one][1])
        twenty_one_cell.value = "%s block： %s" % (conf.twenty_one, str(value[conf.twenty_one]))
        twenty_two_cell = sheet.cell(row=position[conf.twenty_two][0], column=position[conf.twenty_two][1])
        twenty_two_cell.value = "%s block： %s" % (conf.twenty_two, str(value[conf.twenty_two]))
        twenty_three_cell = sheet.cell(row=position[conf.twenty_three][0], column=position[conf.twenty_three][1])
        twenty_three_cell.value = "%s block： %s" % (conf.twenty_three, str(value[conf.twenty_three]))
        twenty_four_cell = sheet.cell(row=position[conf.twenty_four][0], column=position[conf.twenty_four][1])
        twenty_four_cell.value = "%s block： %s" % (conf.twenty_four, str(value[conf.twenty_four]))

        # 只比较21和22的数据
        font_list = [[twenty_one_cell, value[conf.twenty_one]], [twenty_two_cell, value[conf.twenty_two]]]
        self.comparative_white_balance_indicator(font_list)
        wb.save(self.file_path)
        wb.close()

    def write_rgby_noise_data(self, position, value):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]

        r_cell = sheet.cell(row=position[conf.R][0], column=position[conf.R][1])
        r_cell.value = "R：%s%%" % str(value[conf.R])
        g_cell = sheet.cell(row=position[conf.G][0], column=position[conf.G][1])
        g_cell.value = "G：%s%%" % str(value[conf.G])
        b_cell = sheet.cell(row=position[conf.B][0], column=position[conf.B][1])
        b_cell.value = "B：%s%%" % str(value[conf.B])
        y_cell = sheet.cell(row=position[conf.Y][0], column=position[conf.Y][1])
        y_cell.value = "Y：%s%%" % str(value[conf.Y])

        font_list = [[r_cell, value[conf.R]], [g_cell, value[conf.G]], [b_cell, value[conf.B]], [y_cell, value[conf.Y]]]
        self.comparative_rgby_noise_indicator(font_list)

        wb.save(self.file_path)
        wb.close()

    def write_color_accuracy_data(self, position, value):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]

        e1_cell = sheet.cell(row=position[conf.E1][0], column=position[conf.E1][1])
        e1_cell.value = "△E1：%s" % str(value[conf.E1])
        c1_cell = sheet.cell(row=position[conf.C1][0], column=position[conf.C1][1])
        c1_cell.value = "△C1：%s" % str(value[conf.C1])
        c2_cell = sheet.cell(row=position[conf.C2][0], column=position[conf.C2][1])
        c2_cell.value = "△C2：%s" % str(value[conf.C2])
        sat_cell = sheet.cell(row=position[conf.Sat][0], column=position[conf.Sat][1])
        sat_cell.value = "Sat：%s%%" % str("{:.1f}".format(value[conf.Sat]))

        font_data = {conf.E1: [e1_cell, value[conf.E1]], conf.C1: [c1_cell, float(value[conf.C1])],
                     conf.C2: [c2_cell, float(value[conf.C2])], conf.Sat: [sat_cell, float(value[conf.Sat])]}
        self.comparative_saturation_indicator(font_data)
        wb.save(self.file_path)
        wb.close()

    def write_hj_data(self):
        # 灰阶测试
        csv = CSVTestData(conf.hj_data_path)
        hj_data = csv.get_hj_relate_data(csv.read_csv_to_matrix())

        dr_key_pos = self.report_position.find_scenario_position_by_keyword(conf.r_dynamic_range)
        dr_pos = self.report_position.get_hj_relate_position(dr_key_pos)
        self.write_dynamic_range_data(dr_pos, hj_data)

        contrast_key_pos = self.report_position.find_scenario_position_by_keyword(conf.r_contrast_test)
        contrast_pos = self.report_position.get_hj_relate_position(contrast_key_pos)
        self.write_contrast_data(contrast_pos, hj_data)

        readable_key_pos = self.report_position.find_scenario_position_by_keyword(conf.r_readable_hj_num)
        readable_pos = self.report_position.get_hj_relate_position(readable_key_pos)
        self.write_readable_hj_data(readable_pos, hj_data)

    def write_scenario_data(self, light_data_path, scenario):
        self.scenario_light = scenario
        # color check
        csv = CSVTestData(light_data_path)
        data = csv.read_csv_to_matrix()
        # snr
        key_position = self.report_position.find_scenario_position_by_keyword(scenario)
        snr_data = csv.get_snr_data(data)
        snr_position = self.report_position.get_snr_position(key_position)
        self.write_snr_data(snr_position, snr_data)
        # 白平衡
        white_balance_data = csv.get_white_balance_err_data(data)
        white_balance_position = self.report_position.get_white_balance_position(key_position)
        self.write_white_balance_data(white_balance_position, white_balance_data)
        # 噪点测试
        noise_data = csv.get_rgby_noise_data(data)
        noise_position = self.report_position.get_rgby_noise_position(key_position)
        self.write_rgby_noise_data(noise_position, noise_data)
        # 颜色偏差 / 饱和度测试
        saturation_data = csv.get_color_accuracy_data(data)
        saturation_position = self.report_position.get_color_accuracy_position(key_position)
        self.write_color_accuracy_data(saturation_position, saturation_data)

    def write_project_name(self, camera_data):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        position = self.report_position.get_test_project_position(conf.r_test_project)
        r_cell = sheet.cell(row=position[0], column=position[1])
        now = datetime.now()
        if now.month < 10:
            month = "0%d" % now.month
        else:
            month = now.month
        if now.day < 10:
            day = "0%d" % now.day
        else:
            day = now.day
        time_info = "%d%s%s" % (now.year, month, day)
        r_cell.value = "%s-%s万摄像头(%s)-%s" % (
            camera_data["project_name"], str(camera_data["pixels"]), camera_data["camera_product"], time_info)
        wb.save(self.file_path)
        wb.close()

    def comparative_contrast_indicator(self, cell, value):
        if self.is_quality:
            if value < 0.75:
                cell.font = self.red_font
                self.HJ_light_description.append("灰阶可读阶数偏低")
        else:
            if value < 0.7:
                cell.font = self.red_font
                self.HJ_light_description.append("灰阶可读阶数偏低")

    def comparative_readable_hj_indicator(self, cell, value):
        if self.is_quality:
            if value < 15:
                cell.font = self.red_font
                self.HJ_light_description.append("灰阶对比度偏低")
        else:
            if value < 11:
                cell.font = self.red_font
                self.HJ_light_description.append("灰阶对比度偏低")

    def comparative_dynamic_range_indicator(self, cell, value):
        if self.is_quality:
            if value < 5.5:
                cell.font = self.red_font
                self.HJ_light_description.append("灰阶图偏低")
        else:
            if value < 5:
                cell.font = self.red_font
                self.HJ_light_description.append("灰阶图偏低")

    def comparative_snr_indicator(self, cells):
        # l = [["cell", 0], ["cell", 0]]
        for cell in cells:
            if cell[1] < 35:
                cell[0].font = self.red_font
                snr_desc = "信噪比偏低"
                if self.scenario_light == conf.r_F_light:
                    if snr_desc not in self.F_light_description:
                        self.F_light_description.append(snr_desc)
                elif self.scenario_light == conf.r_D65_light:
                    if snr_desc not in self.D65_light_description:
                        self.D65_light_description.append(snr_desc)
                elif self.scenario_light == conf.r_CWF_light:
                    if snr_desc not in self.CWF_light_description:
                        self.CWF_light_description.append(snr_desc)
                else:
                    if snr_desc not in self.TL84_light_description:
                        self.TL84_light_description.append(snr_desc)

    def comparative_white_balance_indicator(self, cells):
        white_balance_des = "白平衡不及格"
        for cell in cells:
            if self.scenario_light == conf.r_F_light:
                if self.is_quality:
                    if cell[1] > 0.3:
                        cell[0].font = self.red_font
                        if white_balance_des not in self.F_light_description:
                            self.F_light_description.append(white_balance_des)
                else:
                    if cell[1] > 0.4:
                        cell[0].font = self.red_font
                        if white_balance_des not in self.F_light_description:
                            self.F_light_description.append(white_balance_des)
            else:
                if self.is_quality:
                    if cell[1] > 0.1:
                        cell[0].font = self.red_font
                        if self.scenario_light == conf.r_D65_light:
                            if white_balance_des not in self.D65_light_description:
                                self.D65_light_description.append(white_balance_des)
                        elif self.scenario_light == conf.r_CWF_light:
                            if white_balance_des not in self.CWF_light_description:
                                self.CWF_light_description.append(white_balance_des)
                        else:
                            if white_balance_des not in self.TL84_light_description:
                                self.TL84_light_description.append(white_balance_des)
                else:
                    if cell[1] > 0.15:
                        cell[0].font = self.red_font
                        if self.scenario_light == conf.r_D65_light:
                            if white_balance_des not in self.D65_light_description:
                                self.D65_light_description.append(white_balance_des)
                        elif self.scenario_light == conf.r_CWF_light:
                            if white_balance_des not in self.CWF_light_description:
                                self.CWF_light_description.append(white_balance_des)
                        else:
                            if white_balance_des not in self.TL84_light_description:
                                self.TL84_light_description.append(white_balance_des)

    def comparative_rgby_noise_indicator(self, cells):
        rgby_noise_des = "噪点偏高"
        for cell in cells:
            if self.is_quality:
                if cell[1] > 2:
                    cell[0].font = self.red_font
                    if self.scenario_light == conf.r_F_light:
                        if rgby_noise_des not in self.F_light_description:
                            self.F_light_description.append(rgby_noise_des)
                    elif self.scenario_light == conf.r_D65_light:
                        if rgby_noise_des not in self.D65_light_description:
                            self.D65_light_description.append(rgby_noise_des)
                    elif self.scenario_light == conf.r_CWF_light:
                        if rgby_noise_des not in self.CWF_light_description:
                            self.CWF_light_description.append(rgby_noise_des)
                    else:
                        if rgby_noise_des not in self.TL84_light_description:
                            self.TL84_light_description.append(rgby_noise_des)
            else:
                if cell[1] > 3:
                    cell[0].font = self.red_font
                    if self.scenario_light == conf.r_F_light:
                        if rgby_noise_des not in self.F_light_description:
                            self.F_light_description.append(rgby_noise_des)
                    elif self.scenario_light == conf.r_D65_light:
                        if rgby_noise_des not in self.D65_light_description:
                            self.D65_light_description.append(rgby_noise_des)
                    elif self.scenario_light == conf.r_CWF_light:
                        if rgby_noise_des not in self.CWF_light_description:
                            self.CWF_light_description.append(rgby_noise_des)
                    else:
                        if rgby_noise_des not in self.TL84_light_description:
                            self.TL84_light_description.append(rgby_noise_des)

    def comparative_saturation_indicator(self, cells):
        E1_cell = cells[conf.E1][0]
        E1_value = cells[conf.E1][1]
        C1_cell = cells[conf.C1][0]
        C1_value = cells[conf.C1][1]
        C2_cell = cells[conf.C2][0]
        C2_value = cells[conf.C2][1]
        Sat_cell = cells[conf.Sat][0]
        Sat_value = cells[conf.Sat][1]
        saturation_des = "颜色还原/饱和度测试不合格"
        flag = 0
        if self.is_quality:
            if E1_value > 15:
                E1_cell.font = self.red_font
                flag += 1

            if self.scenario_light == conf.r_D65_light:
                if C1_value > 10:
                    C1_cell.font = self.red_font
                    flag += 1

                if C2_value > 10:
                    C2_cell.font = self.red_font
                    flag += 1
            else:
                if C1_value > 12:
                    C1_cell.font = self.red_font
                    flag += 1

                if C2_value > 12:
                    C2_cell.font = self.red_font
                    flag += 1

            if self.scenario_light == conf.r_F_light:
                if Sat_value < 90 or Sat_value > 125:
                    Sat_cell.font = self.red_font
                    flag += 1
            else:
                if Sat_value < 100 or Sat_value > 125:
                    Sat_cell.font = self.red_font
                    flag += 1
        else:
            if E1_value > 20:
                E1_cell.font = self.red_font
                flag += 1
            if C1_value > 15:
                C1_cell.font = self.red_font
                flag += 1
            if C2_value > 15:
                C2_cell.font = self.red_font
                flag += 1

            if self.scenario_light == conf.r_F_light:
                if Sat_value < 90 or Sat_value > 125:
                    Sat_cell.font = self.red_font
                    flag += 1
            else:
                if Sat_value < 95 or Sat_value > 135:
                    Sat_cell.font = self.red_font
                    flag += 1

        if flag > 0:
            if self.scenario_light == conf.r_F_light:
                if saturation_des not in self.F_light_description:
                    self.F_light_description.append(saturation_des)
            elif self.scenario_light == conf.r_D65_light:
                if saturation_des not in self.D65_light_description:
                    self.D65_light_description.append(saturation_des)
            elif self.scenario_light == conf.r_CWF_light:
                if saturation_des not in self.CWF_light_description:
                    self.CWF_light_description.append(saturation_des)
            else:
                if saturation_des not in self.TL84_light_description:
                    self.TL84_light_description.append(saturation_des)

    def fill_camera_with_standard(self, standard_param):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        # 摄像头参数位置
        # True的话为精品
        if standard_param:
            first_position = self.report_position.find_scenario_position_by_keyword(conf.r_camera_most_standard)
        else:
            first_position = self.report_position.find_scenario_position_by_keyword(conf.r_camera_standard)

        last_key_position = self.report_position.find_scenario_position_by_keyword(conf.r_camera_p_last_key)

        for i in range(first_position[0], last_key_position[0] + 1):
            cel = sheet.cell(row=i, column=first_position[1])
            cel.fill = self.yellow_fill
        wb.save(self.file_path)
        wb.close()


if __name__ == '__main__':
    w_r = WriteReport(conf.template_path, conf.sheet_name)
    w_r.write_scenario_data(conf.f_data_path, conf.r_F_light)
    w_r.write_hj_data()

    # w_r = WriteReport(conf.template_path, conf.sheet_name)
    # report_position = GetReportPosition(conf.template_path, conf.sheet_name)

    # 灰阶测试
    csv = CSVTestData(conf.hj_data_path)
    hj_data = csv.get_hj_relate_data(csv.read_csv_to_matrix())
    print(hj_data)

    #
    # dr_key_pos = report_position.find_scenario_position_by_keyword(conf.r_dynamic_range)
    # print(dr_key_pos)
    # dr_pos = report_position.get_hj_relate_position(dr_key_pos)
    # print(dr_pos)
    # w_r.write_dynamic_range_data(dr_pos, hj_data)
    #
    # contrast_key_pos = report_position.find_scenario_position_by_keyword(conf.r_contrast_test)
    # print(contrast_key_pos)
    # contrast_pos = report_position.get_hj_relate_position(contrast_key_pos)
    # print(contrast_pos)
    # w_r.write_contrast_data(contrast_pos, hj_data)
    #
    # readable_key_pos = report_position.find_scenario_position_by_keyword(conf.r_readable_hj_num)
    # print(readable_key_pos)
    # readable_pos = report_position.get_hj_relate_position(readable_key_pos)
    # print(readable_pos)
    # w_r.write_readable_hj_data(readable_pos, hj_data)

    # # color check
    # csv = CSVTestData(conf.f_data_path)
    # data = csv.read_csv_to_matrix()
    # # snr
    # f_key_position = report_position.find_scenario_position_by_keyword(conf.r_F_light)
    # snr_data = csv.get_snr_data(data)
    # snr_position = report_position.get_snr_position(f_key_position)
    # w_r.write_snr_data(snr_position, snr_data)
    # # 白平衡
    # white_balance_data = csv.get_white_balance_err_data(data)
    # white_balance_position = report_position.get_white_balance_position(f_key_position)
    # w_r.write_white_balance_data(white_balance_position, white_balance_data)
    # # 噪点测试
    # noise_data = csv.get_rgby_noise_data(data)
    # noise_position = report_position.get_rgby_noise_position(f_key_position)
    # w_r.write_rgby_noise_data(noise_position, noise_data)
    # # 颜色偏差 / 饱和度测试
    # saturation_data = csv.get_color_accuracy_data(data)
    # saturation_position = report_position.get_color_accuracy_position(f_key_position)
    # w_r.write_color_accuracy_data(saturation_position, saturation_data)
