from openpyxl import load_workbook
from openpyxl.styles import Font
from Conf.config import Config
from Common.interface import Interface

conf = Config()


class GetReportPosition(Interface):
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def get_hj_relate_position(self, key_position):
        return {conf.r_hj_relate_pos: [key_position[0], key_position[1] + 1]}

    def find_scenario_position_by_keyword(self, keyword):
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and self.remove_space(keyword) in self.remove_space(cell.value):
                    # 获取单元格的行和列（注意行和列是从1开始的）, 返回x, y
                    wb.close()
                    return [cell.row, cell.column]

    def get_snr_position(self, scenario_position):
        # 顺序分别为RGBY
        column = scenario_position[1] + 1
        r_position = [scenario_position[0] + 1, column]
        g_position = [scenario_position[0] + 2, column]
        b_position = [scenario_position[0] + 3, column]
        y_position = [scenario_position[0] + 4, column]
        return {conf.R: r_position, conf.G: g_position, conf.B: b_position, conf.Y: y_position}

    def get_white_balance_position(self, scenario_position):
        column = scenario_position[1] + 1
        position_19 = [scenario_position[0] + 5, column]
        position_20 = [scenario_position[0] + 6, column]
        position_21 = [scenario_position[0] + 7, column]
        position_22 = [scenario_position[0] + 8, column]
        position_23 = [scenario_position[0] + 9, column]
        position_24 = [scenario_position[0] + 10, column]
        return {conf.nineteen: position_19, conf.twenty: position_20, conf.twenty_one: position_21,
                conf.twenty_two: position_22, conf.twenty_three: position_23, conf.twenty_four: position_24}

    def get_rgby_noise_position(self, scenario_position):
        column = scenario_position[1] + 1
        r_position = [scenario_position[0] + 11, column]
        g_position = [scenario_position[0] + 12, column]
        b_position = [scenario_position[0] + 13, column]
        y_position = [scenario_position[0] + 14, column]
        return {conf.R: r_position, conf.G: g_position, conf.B: b_position, conf.Y: y_position}

    def get_color_accuracy_position(self, scenario_position):
        column = scenario_position[1] + 1
        e1_position = [scenario_position[0] + 15, column]
        c1_position = [scenario_position[0] + 16, column]
        c2_position = [scenario_position[0] + 17, column]
        sat_position = [scenario_position[0] + 18, column]
        return {conf.E1: e1_position, conf.C1: c1_position, conf.C2: c2_position, conf.Sat: sat_position}

    def get_test_project_position(self, key_word):
        key_position = self.find_scenario_position_by_keyword(key_word)
        return [key_position[0], key_position[1] + 1]


if __name__ == '__main__':
    from Common.color_data_filter import CSVTestData

    csv = CSVTestData(conf.hj_data_path)
    hj_data = csv.get_hj_relate_data(csv.read_csv_to_matrix())
    report = GetReportPosition(conf.template_path, conf.sheet_name)
    # report.write_hj_data(hj_data)
    # Dynamic
    scenario = report.find_scenario_position_by_keyword(conf.r_dynamic_range)
    print(scenario)
    print(report.get_hj_relate_position(scenario))
